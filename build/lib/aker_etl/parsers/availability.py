"""Unit Availability parser. Grammar: PLAN.md 0.2.

Every file is exactly 7 rows x 18 columns. Row 6 is the data; row 7 is a "Total"
row that duplicates row 6 and is deliberately not loaded -- on 134c and 139c it
differs from row 6 in pct_trend only (0 vs 100, an upstream rounding artifact on
zero-occupancy books), which is recorded as an `info` issue rather than an error.
"""

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

import openpyxl

from ..models import (
    AvailabilityFile,
    CoercionError,
    ParsedIssue,
    as_decimal,
    as_int,
)

N_COLS = 18

TITLE = "Unit Availability"
HEADER_ROW_1 = ("Property", "Name", "Avg.", "Avg.", "Units", "Occupied", "Vacant", "Vacant",
                "Notice", "Notice", "Avail", "Model", "Down", "Admin", "% Occ", "% Occ",
                "% Leased", "% Trend")
HEADER_ROW_2 = ("", "", "Sq Ft", "Rent", "", "No Notice", "Rented", "Unrented", "Rented",
                "Unrented", "", "", "", "", "", "w/NonRev", "", "")

FIELDS = ["property_code", "property_name", "avg_sqft", "avg_rent", "units",
          "occupied_no_notice", "vacant_rented", "vacant_unrented", "notice_rented",
          "notice_unrented", "available", "model", "down", "admin", "pct_occ",
          "pct_occ_w_nonrev", "pct_leased", "pct_trend"]

_TITLE_RE = re.compile(r"^(?P<name>.+?)\s*\((?P<code>[^)]+)\)$")
_AS_OF_RE = re.compile(r"^As Of\s*=\s*(?P<d>\d{2}/\d{2}/\d{4})$")


class AvailabilityStructureError(RuntimeError):
    pass


def _txt(v: object) -> str:
    return "" if v is None else str(v).strip()


def parse_availability(path: Path) -> AvailabilityFile:
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if wb.sheetnames != ["Report1"]:
            raise AvailabilityStructureError(
                f"{path.name}: expected a single sheet named 'Report1', got {wb.sheetnames}"
            )
        rows: list[list[object]] = []
        for row in wb["Report1"].iter_rows(values_only=True):
            cells = list(row)
            if len(cells) < N_COLS:
                cells.extend([None] * (N_COLS - len(cells)))
            rows.append(cells[:N_COLS])
    finally:
        wb.close()

    if len(rows) < 6:
        raise AvailabilityStructureError(f"{path.name}: {len(rows)} rows, expected at least 6")
    if _txt(rows[0][0]) != TITLE:
        raise AvailabilityStructureError(f"{path.name}: row 1 is {_txt(rows[0][0])!r}, expected {TITLE!r}")

    m = _TITLE_RE.match(_txt(rows[1][0]))
    if not m:
        raise AvailabilityStructureError(
            f"{path.name}: row 2 {_txt(rows[1][0])!r} does not match '<name> (<code>)'"
        )
    property_name, property_code = m.group("name").strip(), m.group("code").strip()

    m_asof = _AS_OF_RE.match(_txt(rows[2][0]))
    if not m_asof:
        raise AvailabilityStructureError(f"{path.name}: row 3 {_txt(rows[2][0])!r} is not 'As Of = MM/DD/YYYY'")
    as_of_date = dt.datetime.strptime(m_asof.group("d"), "%m/%d/%Y").date()

    h1 = tuple(_txt(v) for v in rows[3])
    h2 = tuple(_txt(v) for v in rows[4])
    if h1 != HEADER_ROW_1:
        raise AvailabilityStructureError(f"{path.name}: header row 4 changed.\n got {h1}\nwant {HEADER_ROW_1}")
    if h2 != HEADER_ROW_2:
        raise AvailabilityStructureError(f"{path.name}: header row 5 changed.\n got {h2}\nwant {HEADER_ROW_2}")

    data = rows[5]
    if _txt(data[0]) != property_code:
        raise AvailabilityStructureError(
            f"{path.name}: data row property {_txt(data[0])!r} != title property {property_code!r}"
        )

    issues: list[ParsedIssue] = []

    # Row 7 is the ignored "Total" row. Compare it against row 6 anyway: a
    # difference anywhere other than pct_trend would mean the format changed.
    if len(rows) >= 7:
        total_row = rows[6]
        if _txt(total_row[1]) == "Total":
            for i in range(2, N_COLS):
                if _txt(data[i]) != _txt(total_row[i]):
                    issues.append(
                        ParsedIssue(
                            severity="info" if FIELDS[i] == "pct_trend" else "warning",
                            rule="availability_total_row_differs",
                            sheet_row=7,
                            detail={"property_code": property_code, "field": FIELDS[i],
                                    "data_row": _txt(data[i]), "total_row": _txt(total_row[i])},
                        )
                    )

    def num(i: int) -> int:
        try:
            v = as_int(data[i])
        except CoercionError as exc:
            issues.append(
                ParsedIssue(severity="warning", rule="availability_coercion", sheet_row=6,
                            detail={"property_code": property_code, "field": FIELDS[i],
                                    "error": str(exc)})
            )
            return 0
        return v if v is not None else 0

    return AvailabilityFile(
        property_code=property_code,
        property_name=property_name,
        as_of_date=as_of_date,
        avg_sqft=num(2),
        avg_rent=as_decimal(data[3]),
        units=num(4),
        occupied_no_notice=num(5),
        vacant_rented=num(6),
        vacant_unrented=num(7),
        notice_rented=num(8),
        notice_unrented=num(9),
        available=num(10),
        model=num(11),
        down=num(12),
        admin=num(13),
        pct_occ=as_decimal(data[14]),
        pct_occ_w_nonrev=as_decimal(data[15]),
        pct_leased=as_decimal(data[16]),
        pct_trend=as_decimal(data[17]),
        sheet_rows=len(rows),
        issues=issues,
    )
