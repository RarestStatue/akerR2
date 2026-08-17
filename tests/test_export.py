"""Table export renderers: F7. PLAN6 phase 4. No database, no model."""

from __future__ import annotations

import datetime as dt
import io
from decimal import Decimal

import openpyxl

from aker_etl.dashboard.export import to_csv, to_xlsx


def test_csv_has_a_bom():
    out = to_csv(["a", "b"], [(1, 2)])
    assert out.startswith(b"\xef\xbb\xbf")


def test_xlsx_round_trips_through_openpyxl():
    out = to_xlsx("Sheet1", ["a", "b"], [(1, "x"), (2, "y")])
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("a", "b")
    assert rows[1] == (1, "x")
    assert rows[2] == (2, "y")


def test_decimals_and_dates_stay_typed_in_xlsx():
    out = to_xlsx("Sheet1", ["amount", "as_of"], [(Decimal("12.50"), dt.date(2026, 2, 25))])
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb.active
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert isinstance(row[0], float)
    assert row[0] == 12.5
    assert isinstance(row[1], dt.datetime)
    assert row[1].date() == dt.date(2026, 2, 25)


def test_a_timezone_aware_datetime_does_not_raise():
    # raw.load_issue.created_at is timestamptz; openpyxl rejects tz-aware
    # datetimes outright, so the export must strip the tzinfo, not pass it through.
    aware = dt.datetime(2026, 2, 25, 12, 0, tzinfo=dt.timezone.utc)
    out = to_xlsx("Sheet1", ["created_at"], [(aware,)])
    wb = openpyxl.load_workbook(io.BytesIO(out))
    row = list(wb.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row[0] == dt.datetime(2026, 2, 25, 12, 0)


def test_jsonb_and_array_columns_are_stringified():
    out = to_xlsx(
        "Sheet1", ["detail", "codes"],
        [({"rule": "x", "n": 3}, ["115r", "126a"])],
    )
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb.active
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert isinstance(row[0], str) and "rule" in row[0]
    assert isinstance(row[1], str) and "115r" in row[1]


def test_the_header_is_frozen_and_columns_are_sized():
    # write_only sheets emit sheetViews and <cols> on the first append, so this
    # is the assertion that catches them being set too late.
    import io

    import openpyxl

    from aker_etl.dashboard.export import to_xlsx

    wb = openpyxl.load_workbook(
        io.BytesIO(to_xlsx("t", ["alpha", "b"], [["a very long value indeed here", 1], ["x", 2]]))
    )
    ws = wb.active
    assert ws.freeze_panes == "A2"
    assert ws.column_dimensions["A"].width == 31.0
    assert ws.column_dimensions["B"].width == 8.0
