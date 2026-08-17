"""Table export. One registry, so a dataset cannot exist in the UI and not here.

Every entry is a parameterised SELECT taking exactly one parameter, the
snapshot_id. Nothing accepts SQL from the caller: the dataset name indexes this
dict and an unknown name is a 404, which is what keeps a query-string parameter
from becoming a query.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DATASETS: dict[str, tuple[str, str]] = {
    # name: (filename stem, SQL taking one %s = snapshot_id)
    "properties": (
        "properties",
        """SELECT p.property_code::text AS property_code, p.property_name,
                  p.book_type::text AS book_type, p.asset_key,
                  COALESCE(k.units, 0) AS units, COALESCE(k.occupied_units, 0) AS occupied_units,
                  COALESCE(k.notice_units, 0) AS notice_units,
                  COALESCE(k.vacant_units, 0) AS vacant_units,
                  COALESCE(k.non_revenue_units, 0) AS non_revenue_units,
                  COALESCE(k.future_leases, 0) AS future_leases,
                  k.pct_occupied, COALESCE(k.market_rent, 0) AS market_rent,
                  COALESCE(k.lease_charges, 0) AS lease_charges,
                  COALESCE(k.square_feet, 0) AS square_feet, COALESCE(k.balance, 0) AS balance
           FROM core.property p
           LEFT JOIN mart.property_snapshot_kpi k
             ON k.property_id = p.property_id AND k.snapshot_id = %s
           ORDER BY p.property_code""",
    ),
    "assets": (
        "assets",
        """SELECT asset_key, asset_label, book_count, property_codes, units, occupied_units,
                  vacant_units, pct_occupied, market_rent, lease_charges, balance
           FROM mart.asset_snapshot_kpi WHERE snapshot_id = %s ORDER BY units DESC""",
    ),
    # Deliberately unpaginated -- exporting page 1 of 21 is not an export.
    "units": (
        "units",
        """SELECT p.property_code::text AS property_code, u.unit_code, ut.unit_type_code,
                  l.unit_sqft, l.occupancy_status::text AS occupancy_status,
                  l.section::text AS section, l.resident_id, r.display_name,
                  l.market_rent, l.charges_total, l.balance, l.move_in, l.lease_expiration,
                  l.move_out
           FROM core.lease l
           JOIN core.property p ON p.property_id = l.property_id
           JOIN core.unit u ON u.unit_id = l.unit_id
           LEFT JOIN core.unit_type ut ON ut.unit_type_id = l.unit_type_id
           LEFT JOIN core.resident r ON r.resident_id = l.resident_id
           WHERE l.snapshot_id = %s
           ORDER BY p.property_code, u.unit_code, l.section""",
    ),
    "charge_mix": (
        "charge-mix",
        """SELECT p.property_code::text AS property_code, cm.category::text AS category,
                  cm.charge_code, cm.line_count, cm.amount
           FROM mart.charge_mix cm JOIN core.property p USING (property_id)
           WHERE cm.snapshot_id = %s ORDER BY p.property_code, cm.category, cm.charge_code""",
    ),
    "expirations": (
        "expirations",
        """SELECT p.property_code::text AS property_code, e.expiry_month::text AS expiry_month,
                  e.expiring_leases, e.charges_at_risk, e.holdover_mtm
           FROM mart.expiration_schedule e JOIN core.property p USING (property_id)
           WHERE e.snapshot_id = %s ORDER BY p.property_code, e.expiry_month""",
    ),
    "matrix": (
        "profitability-matrix",
        """SELECT property_code::text AS property_code, property_name,
                  book_type::text AS book_type, asset_key, units, occupied_units,
                  notice_units, vacant_units, pct_occupied, market_rent, lease_charges,
                  revenue_capture_pct, quadrant, plottable, exclusion_reason,
                  charge_coverage, charges_to_threshold, units_to_threshold,
                  loss_to_lease, concessions, ancillary_charges, units_owing, balance_owed
           FROM mart.property_profitability WHERE snapshot_id = %s ORDER BY property_code""",
    ),
    "revenue_bridge": (
        "revenue-bridge",
        """SELECT property_code, property_name, book_type, asset_key, gross_potential_rent,
                  vacancy_loss, loss_to_lease, rent_charges, subsidy, concessions, ancillary,
                  billed_charges, charge_coverage, exclusion_reason
           FROM mart.revenue_bridge WHERE snapshot_id = %s ORDER BY property_code""",
    ),
    "outliers": (
        "rent-positioning",
        """SELECT property_code, unit_code, unit_type_code, unit_sqft, occupancy_status,
                  lease_expiration, market_rent, contract_rent, rent_psf, peer_units,
                  median_market_rent, median_rent_psf, market_vs_median, pct_vs_median,
                  contract_vs_market
           FROM mart.unit_rent_outlier WHERE snapshot_id = %s
           ORDER BY property_code, unit_code""",
    ),
    "concentration": (
        "expiration-concentration",
        """SELECT property_code, property_name, expiry_month::text AS expiry_month,
                  expiring_leases, dated_leases, share_of_book, concentrated, leases_to_shift,
                  charges_at_risk, holdover_mtm
           FROM mart.expiration_concentration WHERE snapshot_id = %s
           ORDER BY property_code, expiry_month""",
    ),
    "anomalies": (
        "computed-findings",
        """SELECT property_code, property_name, units, metric, label, unit, worse_when,
                  value, peer_mean, peer_sd, peer_books, z, adverse, priority
           FROM mart.property_anomaly WHERE snapshot_id = %s
           ORDER BY property_code, metric""",
    ),
    "reconciliation": (
        "reconciliation",
        """SELECT property_code, detail_units, availability_units, unit_delta,
                  detail_occupied, availability_occupied, detail_charges, report_charges,
                  charge_delta
           FROM mart.reconciliation WHERE snapshot_id = %s ORDER BY property_code""",
    ),
    "issues": (
        "load-issues",
        """SELECT severity::text AS severity, rule, sheet_row, detail, created_at
           FROM raw.load_issue
           WHERE run_id IN (SELECT DISTINCT sf.run_id
                            FROM core.lease l
                            JOIN raw.source_file sf ON sf.file_id = l.file_id
                            WHERE l.snapshot_id = %s)
           ORDER BY severity, rule""",
    ),
    "charge_codes": (
        "charge-codes",
        """SELECT cc.charge_code, cc.category::text AS category, cc.description,
                  cc.is_concession, cc.label_verified,
                  count(lc.*) FILTER (WHERE l.lease_id IS NOT NULL)        AS line_count,
                  coalesce(sum(lc.amount) FILTER (WHERE l.lease_id IS NOT NULL), 0) AS amount,
                  count(DISTINCT l.property_id)                            AS properties
           FROM core.charge_code cc
           LEFT JOIN core.lease_charge lc ON lc.charge_code = cc.charge_code
           LEFT JOIN core.lease l         ON l.lease_id = lc.lease_id AND l.snapshot_id = %s
           GROUP BY cc.charge_code, cc.category, cc.description, cc.is_concession,
                    cc.label_verified
           ORDER BY cc.category, cc.charge_code""",
    ),
}


def to_csv(cols: list[str], rows: list[tuple]) -> bytes:
    """utf-8-sig, not utf-8: Excel on Windows reads a plain utf-8 CSV as cp1252
    and turns every non-ASCII property name into mojibake. The BOM is what makes
    a double-click work."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(cols)
    for row in rows:
        writer.writerow(["" if v is None else str(v) for v in row])
    return buf.getvalue().encode("utf-8-sig")


def _xlsx_cell(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, dt.datetime):
        # Excel has no timezone concept; a tz-aware value (e.g. raw.load_issue's
        # timestamptz columns) makes openpyxl raise. Naive local-clock time is
        # what a spreadsheet can actually show.
        return value.replace(tzinfo=None) if value.tzinfo is not None else value
    if isinstance(value, (dt.date, int, float, str)) or value is None:
        return value
    if isinstance(value, bool):
        return value
    return str(value)


def to_xlsx(sheet: str, cols: list[str], rows: list[tuple]) -> bytes:
    """openpyxl, write_only=True. Bold header row frozen at A2, column widths
    from the longest cell capped at 60. Decimals and dates go in as numbers and
    dates, not strings, or the recipient's first act is a text-to-columns.

    Widths and freeze_panes are set BEFORE the first append: a write-only sheet
    streams its XML, and `sheetViews` and `<cols>` are both emitted when the
    first row arrives. Setting either afterwards is discarded silently, which is
    why the widths are computed in their own pass rather than accumulated while
    appending.
    """
    typed_rows = [[_xlsx_cell(v) for v in row] for row in rows]

    widths = [len(c) for c in cols]
    for row in typed_rows:
        for i, v in enumerate(row):
            if i < len(widths):
                widths[i] = min(60, max(widths[i], len(str(v)) if v is not None else 0))

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet[:31] or "Sheet1")
    ws.freeze_panes = "A2"
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(60, max(8, w + 2))

    header_font = Font(bold=True)
    header_cells = []
    for c in cols:
        cell = WriteOnlyCell(ws, value=c)
        cell.font = header_font
        header_cells.append(cell)
    ws.append(header_cells)
    for row in typed_rows:
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
