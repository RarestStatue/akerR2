"""Rent Roll with Lease Charges parser. Grammar: PLAN.md 0.1.

The classifier below is order-dependent and the order is load-bearing. The two
section labels also appear as *summary-group* labels inside the "Summary Groups"
block, so a label test alone injects four fake leases per file. Two things stop
that: an explicit mode flag that is authoritative, and the rule-3 requirement
that B..N be blank on a real section header.
"""

from __future__ import annotations

import datetime as dt
import re
from decimal import Decimal
from pathlib import Path

import openpyxl

from ..models import (
    CoercionError,
    ParsedCharge,
    ParsedChargeSummary,
    ParsedIssue,
    ParsedLease,
    ParsedSummaryGroup,
    RentRollFile,
    as_date,
    as_decimal,
    as_int,
    as_text,
    derive_status,
)

N_COLS = 14
TOTAL_TOLERANCE = Decimal("0.005")

TITLE = "Rent Roll with Lease Charges"
HEADER_ROW_1 = ("Unit", "Unit Type", "Unit", "Resident", "Name", "Market", "Charge",
                "Amount", "Resident", "Other", "Move In", "Lease", "Move Out", "Balance")
HEADER_ROW_2 = ("", "", "Sq Ft", "", "", "Rent", "Code", "", "Deposit", "Deposit",
                "", "Expiration", "", "")

SECTION_LABELS = {
    "Current/Notice/Vacant Residents": "current",
    "Future Residents/Applicants": "future",
}
SUMMARY_GROUP_LABELS = {
    "Current/Notice/Vacant Residents",
    "Future Residents/Applicants",
    "Occupied Units",
    "Total Non Rev Units",
    "Total Vacant Units",
    "Totals:",
}
SENTINELS = {"VACANT", "MODEL", "DOWN"}

_TITLE_RE = re.compile(r"^(?P<name>.+?)\s*\((?P<code>[^)]+)\)$")
_AS_OF_RE = re.compile(r"^As Of\s*=\s*(?P<d>\d{2}/\d{2}/\d{4})$")
_MONTH_RE = re.compile(r"^Month Year\s*=\s*(?P<m>\d{2})/(?P<y>\d{4})$")


class RentRollStructureError(RuntimeError):
    """Structural failure: wrong sheet, wrong header, unparseable title.

    Reserved for things that mean the report format changed. A single bad *row*
    never raises -- it becomes a ParsedIssue and the load continues.
    """


def _txt(v: object) -> str:
    return "" if v is None else str(v).strip()


def _read_rows(path: Path) -> list[list[object]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if wb.sheetnames != ["Report1"]:
            raise RentRollStructureError(
                f"{path.name}: expected a single sheet named 'Report1', got {wb.sheetnames}"
            )
        ws = wb["Report1"]
        rows: list[list[object]] = []
        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            # openpyxl truncates trailing empty cells, so rows come back short.
            # Right-padding here is what keeps every downstream index safe.
            if len(cells) < N_COLS:
                cells.extend([None] * (N_COLS - len(cells)))
            rows.append(cells[:N_COLS])
        return rows
    finally:
        wb.close()


def parse_rent_roll(path: Path) -> RentRollFile:
    path = Path(path)
    rows = _read_rows(path)
    if len(rows) < 6:
        raise RentRollStructureError(f"{path.name}: only {len(rows)} rows, expected a header block")

    if _txt(rows[0][0]) != TITLE:
        raise RentRollStructureError(f"{path.name}: row 1 is {_txt(rows[0][0])!r}, expected {TITLE!r}")

    m = _TITLE_RE.match(_txt(rows[1][0]))
    if not m:
        raise RentRollStructureError(
            f"{path.name}: row 2 {_txt(rows[1][0])!r} does not match '<name> (<code>)'"
        )
    property_name, property_code = m.group("name").strip(), m.group("code").strip()

    m_asof = _AS_OF_RE.match(_txt(rows[2][0]))
    if not m_asof:
        raise RentRollStructureError(f"{path.name}: row 3 {_txt(rows[2][0])!r} is not 'As Of = MM/DD/YYYY'")
    as_of_date = dt.datetime.strptime(m_asof.group("d"), "%m/%d/%Y").date()

    m_month = _MONTH_RE.match(_txt(rows[3][0]))
    if not m_month:
        raise RentRollStructureError(f"{path.name}: row 4 {_txt(rows[3][0])!r} is not 'Month Year = MM/YYYY'")
    report_month = dt.date(int(m_month.group("y")), int(m_month.group("m")), 1)

    h1 = tuple(_txt(v) for v in rows[4])
    h2 = tuple(_txt(v) for v in rows[5])
    if h1 != HEADER_ROW_1:
        raise RentRollStructureError(f"{path.name}: header row 5 changed.\n got {h1}\nwant {HEADER_ROW_1}")
    if h2 != HEADER_ROW_2:
        raise RentRollStructureError(f"{path.name}: header row 6 changed.\n got {h2}\nwant {HEADER_ROW_2}")

    out = RentRollFile(
        property_code=property_code,
        property_name=property_name,
        as_of_date=as_of_date,
        report_month=report_month,
        sheet_rows=len(rows),
    )

    mode: str | None = None      # None | 'summary_groups' | 'charge_summary'
    skip = 0
    section: str | None = None
    current: ParsedLease | None = None
    charge_sum = Decimal(0)      # running sum for the open block

    def close_block(lease: ParsedLease | None, total: Decimal) -> None:
        """Reconcile the block's printed Total against the sum of its charge lines."""
        if lease is None:
            return
        if abs(lease.charges_total - total) > TOTAL_TOLERANCE:
            out.issues.append(
                ParsedIssue(
                    severity="error",
                    rule="block_total_mismatch",
                    sheet_row=lease.sheet_row,
                    detail={
                        "property_code": property_code,
                        "unit_code": lease.unit_code,
                        "printed_total": str(lease.charges_total),
                        "sum_of_charges": str(total),
                    },
                )
            )

    for idx, r in enumerate(rows[6:], start=7):   # sheet_row is 1-based
        if skip:
            skip -= 1
            continue

        a, d, g = _txt(r[0]), _txt(r[3]), _txt(r[6])

        # 1 / 2 -- mode switches. A summary block always terminates the open lease.
        if a == "Summary Groups":
            close_block(current, charge_sum)
            current, charge_sum = None, Decimal(0)
            mode, skip = "summary_groups", 1       # skip the 2-line header
            continue
        if a.startswith("Summary of Charges"):
            close_block(current, charge_sum)
            current, charge_sum = None, Decimal(0)
            mode, skip = "charge_summary", 2       # '(Current/Notice...)' + 'Charge Code|Amount'
            continue

        if mode == "summary_groups":
            if a == "":
                mode = None
                continue
            if a in SUMMARY_GROUP_LABELS:
                out.summary_groups.append(_parse_summary_group(r, idx, out.issues, property_code))
                continue
            # Unrecognised label inside the block: record it, leave the mode.
            out.issues.append(
                ParsedIssue(severity="warning", rule="unexpected_summary_group_label",
                            sheet_row=idx, detail={"property_code": property_code, "label": a})
            )
            mode = None
            continue

        if mode == "charge_summary":
            if a in ("", "Total"):
                mode = None
                continue
            try:
                out.charge_summary.append(
                    ParsedChargeSummary(charge_code=a, amount=as_decimal(r[3]), sheet_row=idx)
                )
            except CoercionError as exc:
                out.issues.append(
                    ParsedIssue(severity="error", rule="charge_summary_coercion", sheet_row=idx,
                                detail={"property_code": property_code, "error": str(exc)})
                )
            continue

        # 3 -- section switch. The B..N blank test is what disambiguates this from
        # the identically-labelled summary-group rows.
        if a in SECTION_LABELS and all(_txt(v) == "" for v in r[1:]):
            close_block(current, charge_sum)
            current, charge_sum = None, Decimal(0)
            section = SECTION_LABELS[a]
            continue

        # 4 -- new lease block
        if a != "":
            close_block(current, charge_sum)
            current, charge_sum = None, Decimal(0)
            if section is None:
                out.issues.append(
                    ParsedIssue(severity="error", rule="lease_outside_section", sheet_row=idx,
                                detail={"property_code": property_code, "unit_code": a})
                )
                continue
            lease = _parse_lease(r, idx, section, out.issues, property_code)
            if lease is None:
                continue
            current = lease
            out.leases.append(lease)
            if g:
                charge = _parse_charge(g, r[7], 1, idx, out.issues, property_code)
                if charge:
                    lease.charges.append(charge)
                    charge_sum += charge.amount
            continue

        # 5 -- charge line
        if g and g != "Total":
            if current is None:
                out.issues.append(
                    ParsedIssue(severity="error", rule="charge_without_block", sheet_row=idx,
                                detail={"property_code": property_code, "charge_code": g})
                )
                continue
            charge = _parse_charge(g, r[7], len(current.charges) + 1, idx, out.issues, property_code)
            if charge:
                current.charges.append(charge)
                charge_sum += charge.amount
            continue

        # 6 -- the block's Total row
        if g == "Total":
            if current is None:
                continue
            try:
                current.charges_total = as_decimal(r[7])
            except CoercionError as exc:
                out.issues.append(
                    ParsedIssue(severity="error", rule="block_total_coercion", sheet_row=idx,
                                detail={"property_code": property_code, "error": str(exc)})
                )
            close_block(current, charge_sum)
            current, charge_sum = None, Decimal(0)
            continue

        # 7 -- section grand total: closes the section, is not loaded
        if d == "Total":
            close_block(current, charge_sum)
            current, charge_sum = None, Decimal(0)
            continue

        # 8 -- separator
        continue

    close_block(current, charge_sum)
    return out


def _parse_charge(
    code: str, amount: object, line_no: int, sheet_row: int,
    issues: list[ParsedIssue], property_code: str,
) -> ParsedCharge | None:
    try:
        return ParsedCharge(
            line_no=line_no, charge_code=code, amount=as_decimal(amount), sheet_row=sheet_row
        )
    except CoercionError as exc:
        issues.append(
            ParsedIssue(severity="error", rule="charge_coercion", sheet_row=sheet_row,
                        detail={"property_code": property_code, "charge_code": code,
                                "error": str(exc)})
        )
        return None


def _parse_lease(
    r: list[object], sheet_row: int, section: str,
    issues: list[ParsedIssue], property_code: str,
) -> ParsedLease | None:
    """Column map: PLAN.md 0.1. Never raises -- a bad row becomes an issue."""
    unit_code = _txt(r[0])
    try:
        raw_resident = _txt(r[3])
        sentinel = raw_resident if raw_resident in SENTINELS else None
        resident_id = None if sentinel else (raw_resident or None)
        move_out = as_date(r[12])
        status = derive_status(section, sentinel, move_out)  # type: ignore[arg-type]
        lease = ParsedLease(
            unit_code=unit_code,
            unit_type_code=as_text(r[1]),
            unit_sqft=as_int(r[2]),
            resident_id=resident_id,
            resident_name=as_text(r[4]),
            sentinel=sentinel,
            section=section,                                  # type: ignore[arg-type]
            occupancy_status=status,
            market_rent=as_decimal(r[5]),
            resident_deposit=as_decimal(r[8]),
            other_deposit=as_decimal(r[9]),
            balance=as_decimal(r[13]),
            charges_total=Decimal(0),
            move_in=as_date(r[10]),
            lease_expiration=as_date(r[11]),
            move_out=move_out,
            sheet_row=sheet_row,
        )
    except CoercionError as exc:
        issues.append(
            ParsedIssue(severity="error", rule="lease_row_coercion", sheet_row=sheet_row,
                        detail={"property_code": property_code, "unit_code": unit_code,
                                "error": str(exc)})
        )
        return None

    # Sentinels are supposed to carry no dates and a zero balance. A violation is
    # real data worth seeing, not a reason to reject the row.
    if lease.sentinel and (lease.move_in or lease.lease_expiration or lease.move_out
                           or lease.balance != 0):
        issues.append(
            ParsedIssue(severity="warning", rule="sentinel_field_violation", sheet_row=sheet_row,
                        detail={"property_code": property_code, "unit_code": unit_code,
                                "sentinel": lease.sentinel, "balance": str(lease.balance)})
        )
    if resident_id and not re.match(r"^t[A-Za-z0-9]+$", resident_id):
        issues.append(
            ParsedIssue(severity="warning", rule="resident_id_format", sheet_row=sheet_row,
                        detail={"property_code": property_code, "resident_id": resident_id})
        )
    return lease


def _parse_summary_group(
    r: list[object], sheet_row: int, issues: list[ParsedIssue], property_code: str,
) -> ParsedSummaryGroup:
    """Money here arrives as comma-formatted strings ('260,778.00'); same coercers."""

    def money(v: object) -> Decimal | None:
        if v is None or _txt(v) == "":
            return None
        try:
            return as_decimal(v)
        except CoercionError as exc:
            issues.append(
                ParsedIssue(severity="warning", rule="summary_group_coercion", sheet_row=sheet_row,
                            detail={"property_code": property_code, "error": str(exc)})
            )
            return None

    def count(v: object) -> int | None:
        try:
            return as_int(v)
        except CoercionError as exc:
            issues.append(
                ParsedIssue(severity="warning", rule="summary_group_coercion", sheet_row=sheet_row,
                            detail={"property_code": property_code, "error": str(exc)})
            )
            return None

    return ParsedSummaryGroup(
        group_label=_txt(r[0]),
        square_footage=money(r[5]),
        market_rent=money(r[6]),
        lease_charges=money(r[7]),
        security_deposit=money(r[8]),
        other_deposits=money(r[9]),
        unit_count=count(r[10]),
        pct_unit_occupancy=money(r[11]),
        pct_sqft_occupied=money(r[12]),
        balance=money(r[13]),
        sheet_row=sheet_row,
    )
